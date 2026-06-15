"""
NSUT NCC — Flask Backend
Serves the React SPA, provides REST API for attendance, camps, auth, admin.
"""

from flask import (
    Flask, request, redirect, url_for, jsonify,
    send_file, send_from_directory, session
)
import mysql.connector
import pandas as pd
import io
import os
import re
import json
import uuid
import secrets
import hashlib
import tempfile
import smtplib
from functools import wraps
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash

from config import Config

# ── App Setup ─────────────────────────────────────────────────
app = Flask(__name__, static_folder='../frontend/dist', static_url_path='')
app.secret_key = Config.SECRET_KEY
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)

# ── Constants ─────────────────────────────────────────────────
PHOTO_ALLOWED = {'.jpg', '.jpeg', '.png', '.webp', '.gif', '.heic'}
IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.webp', '.gif'}

RANK_PRIORITY = {
    'suo': 1, 'juo': 2, 'csm': 3, 'chm': 3, 'cqms': 4,
    'sergeant': 5, 'sgt': 5, 'corporal': 6, 'cpl': 6,
    'lcpl': 7, 'lance corporal': 7, 'cadet': 8, 'cdt': 8,
}


# ── Helpers ───────────────────────────────────────────────────

def get_db_connection():
    """Create and return a MySQL connection."""
    return mysql.connector.connect(
        host=Config.DB_HOST,
        user=Config.DB_USER,
        password=Config.DB_PASSWORD,
        database=Config.DB_NAME,
    )


def get_session_photos_dir(session_id):
    base = os.path.join(app.root_path, 'static', 'images', 'session_photos', str(session_id))
    os.makedirs(base, exist_ok=True)
    return base


def get_camps_dir():
    return os.path.join(app.root_path, 'static', 'images', 'camps')


def get_achievements_dir():
    base = os.path.join(app.root_path, 'static', 'images', 'achievements')
    os.makedirs(base, exist_ok=True)
    return base


def list_images(folder_path):
    try:
        files = os.listdir(folder_path)
    except FileNotFoundError:
        return []
    return sorted(
        f for f in files
        if os.path.splitext(f)[1].lower() in IMAGE_EXTS
    )


def get_rank_priority(rank_val):
    if not rank_val:
        return 99
    key = str(rank_val).strip().lower()
    return RANK_PRIORITY.get(key, 50)


def extract_dli_year(dli_val):
    if not dli_val:
        return 9999
    match = re.search(r'\d{4}', str(dli_val))
    return int(match.group()) if match else 9999


def validate_email(email):
    """Basic email format validation."""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email))


def validate_dli(dli):
    """Validate DLI number format — at least 6 chars, alphanumeric."""
    if not dli or len(dli.strip()) < 6:
        return False
    return bool(re.match(r'^[A-Za-z0-9]+$', dli.strip()))


def send_reset_email(to_email, reset_code):
    """Send password reset code via email."""
    if not Config.MAIL_USERNAME or not Config.MAIL_PASSWORD:
        print(f"[DEV MODE] Reset code for {to_email}: {reset_code}")
        return True

    try:
        msg = MIMEMultipart()
        msg['From'] = Config.MAIL_DEFAULT_SENDER or Config.MAIL_USERNAME
        msg['To'] = to_email
        msg['Subject'] = 'NSUT NCC — Password Reset Code'

        body = f"""
        <html>
        <body style="font-family: 'Segoe UI', sans-serif; background: #0e1a0e; color: #f5f0e0; padding: 20px;">
            <div style="max-width: 480px; margin: 0 auto; padding: 30px; background: rgba(30,58,30,0.8); border: 1px solid rgba(201,168,76,0.3); border-radius: 12px;">
                <h2 style="color: #c9a84c; margin-bottom: 10px;">NSUT NCC Defense Hub</h2>
                <p>Your password reset code is:</p>
                <div style="background: rgba(201,168,76,0.15); border: 1px solid rgba(201,168,76,0.4); padding: 16px; border-radius: 8px; text-align: center; margin: 20px 0;">
                    <span style="font-size: 28px; font-weight: bold; letter-spacing: 6px; color: #f0c040;">{reset_code}</span>
                </div>
                <p style="font-size: 13px; color: #7a8870;">This code expires in {Config.RESET_TOKEN_EXPIRY_MINUTES} minutes. If you didn't request this, ignore this email.</p>
            </div>
        </body>
        </html>
        """
        msg.attach(MIMEText(body, 'html'))

        server = smtplib.SMTP(Config.MAIL_SERVER, Config.MAIL_PORT)
        server.starttls()
        server.login(Config.MAIL_USERNAME, Config.MAIL_PASSWORD)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"[MAIL ERROR] {e}")
        return False


# ── Auth Decorators ───────────────────────────────────────────

def login_required(f):
    """Ensure user is authenticated."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Authentication required'}), 401
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    """Ensure user is authenticated AND is admin."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Authentication required'}), 401
        if not session.get('is_admin'):
            return jsonify({'error': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return decorated


# ── DB Auto-Migration ─────────────────────────────────────────

def ensure_tables():
    """Create auth tables if they don't exist."""
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INT AUTO_INCREMENT PRIMARY KEY,
                email VARCHAR(255) NOT NULL,
                dli_number VARCHAR(100) NOT NULL,
                password_hash VARCHAR(255) NOT NULL,
                is_admin TINYINT(1) NOT NULL DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY uq_users_email (email),
                UNIQUE KEY uq_users_dli (dli_number)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS password_reset_tokens (
                id INT AUTO_INCREMENT PRIMARY KEY,
                user_id INT NOT NULL,
                token_hash VARCHAR(255) NOT NULL,
                expires_at DATETIME NOT NULL,
                used TINYINT(1) NOT NULL DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS achievements (
                id INT AUTO_INCREMENT PRIMARY KEY,
                title VARCHAR(255) NOT NULL,
                description TEXT NOT NULL,
                image_path VARCHAR(255) NOT NULL,
                tag VARCHAR(100) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        
        cursor.execute("SELECT COUNT(*) FROM achievements")
        count = cursor.fetchone()[0]
        if count <= 3:
            if count > 0:
                cursor.execute("DELETE FROM achievements")
            initial_achievements = [
                ('Harshit Tehlan — Completed BMC',
                 'Completion of BMC reflects disciplined training, steady resolve, and exemplary turnout during the course. A proud milestone for NSUT NCC.',
                 '/static/images/camps/snic/3.png',
                 'BMC COMPLETION'),
                ('Tanishq Sir — Recommended',
                 'A recommendation is earned through consistent performance, mentorship, and leadership qualities. NSUT NCC salutes this honorable recognition.',
                 '/static/images/camps/snic/7.png',
                 'RECOMMENDED'),
                ('Sumit — Became the SUO',
                 'Assuming the SUO role demonstrates command confidence, responsibility, and institutional pride. Congratulations to a leader in formation.',
                 '/static/images/camps/uptrek/IMG-20250414-WA0128.jpg',
                 'SUO APPOINTMENT'),
                ('Placeholder — Annual Training Camp Excellence',
                 'Details coming soon. This placeholder represents future achievement content to be added.',
                 '/static/images/camps/snic/1.jpg',
                 'ATC EXCELLENCE'),
                ('Placeholder — Republic Day Camp Representation',
                 'Details coming soon. This placeholder represents future achievement content to be added.',
                 '/static/images/camps/uptrek/IMG-20250414-WA0128.jpg',
                 'RDC NOMINATION'),
                ('Placeholder — Best Cadet Award',
                 'Details coming soon. This placeholder represents future achievement content to be added.',
                 '/static/images/camps/snic/5.jpg',
                 'BEST CADET'),
                ('Cadet Aditya — Gold Medal in Shooting',
                 'Demonstrated exceptional marksmanship at the inter-battalion shooting competition, securing the top position with precision and composure.',
                 '/static/images/camps/snic/4.jpg',
                 'SHOOTING GOLD'),
                ('Cadet Priya — Selected for TSC',
                 'TSC selection is a testament to unwavering dedication, physical endurance, and leadership potential. A proud achievement for the unit.',
                 '/static/images/camps/uptrek/IMG-20250414-WA0129.jpg',
                 'TSC SELECTION'),
                ('NSUT NCC — Best Marching Contingent',
                 'The unit secured the Best Marching Contingent award at the annual NCC day parade, showcasing synchronized drill and disciplined bearing.',
                 '/static/images/camps/snic/3.png',
                 'BEST CONTINGENT'),
                ('Cadet Rohan — Commander\'s Medal',
                 'Awarded the Commander\'s Medal for outstanding performance across all training domains including drill, campcraft, and community service.',
                 '/static/images/camps/snic/7.png',
                 'COMMANDER MEDAL'),
                ('Cadet Simran — CATC Excellence Trophy',
                 'Won the CATC Overall Excellence Trophy for exemplary conduct, leadership, and skill demonstration during the combined annual training camp.',
                 '/static/images/camps/uptrek/IMG-20250414-WA0128.jpg',
                 'CATC EXCELLENCE'),
                ('NSUT NCC — Clean Campus Drive Award',
                 'Recognized for organizing and executing a highly impactful campus cleanliness and awareness drive under the Swachh Bharat Abhiyan initiative.',
                 '/static/images/camps/snic/1.jpg',
                 'SOCIAL SERVICE')
            ]
            cursor.executemany(
                "INSERT INTO achievements (title, description, image_path, tag) VALUES (%s, %s, %s, %s)",
                initial_achievements
            )
        conn.commit()
    except Exception as e:
        print(f"[DB MIGRATION] {e}")
    finally:
        cursor.close()
        conn.close()


# ── Static File Routes ────────────────────────────────────────

@app.route('/static/images/<path:filename>')
def serve_static_images(filename):
    return send_from_directory(
        os.path.join(app.root_path, 'static', 'images'), filename
    )


# ══════════════════════════════════════════════════════════════
#   AUTH ROUTES
# ══════════════════════════════════════════════════════════════

@app.route('/auth/signup', methods=['POST'])
def auth_signup():
    """Register a new cadet/admin user."""
    data = request.get_json() or {}
    email = (data.get('email') or '').strip().lower()
    dli = (data.get('dli_number') or '').strip().upper()
    password = data.get('password') or ''
    confirm = data.get('confirm_password') or ''

    # Validate
    errors = []
    if not email or not validate_email(email):
        errors.append('Valid email address is required.')
    if not dli or not validate_dli(dli):
        errors.append('Valid DLI number is required (minimum 6 alphanumeric characters).')
    if not password or len(password) < 6:
        errors.append('Password must be at least 6 characters.')
    if password != confirm:
        errors.append('Passwords do not match.')
    if errors:
        return jsonify({'error': ' '.join(errors)}), 400

    # Check duplicates
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            "SELECT id FROM users WHERE email = %s OR dli_number = %s",
            (email, dli)
        )
        existing = cursor.fetchone()
        if existing:
            cursor.close()
            conn.close()
            return jsonify({'error': 'An account with this email or DLI number already exists.'}), 409

        # Hash password and insert
        pw_hash = generate_password_hash(password)
        is_admin = 1 if dli == Config.ADMIN_DLI else 0

        cursor.execute(
            "INSERT INTO users (email, dli_number, password_hash, is_admin) VALUES (%s, %s, %s, %s)",
            (email, dli, pw_hash, is_admin)
        )
        conn.commit()
        user_id = cursor.lastrowid

        # Auto-login after signup
        session.permanent = True
        session['user_id'] = user_id
        session['email'] = email
        session['dli_number'] = dli
        session['is_admin'] = bool(is_admin)

        # Fetch cadet details if exists
        cursor.execute("SELECT name, _rank FROM cadets WHERE dli = %s LIMIT 1", (dli,))
        cadet = cursor.fetchone()
        rank_name = cadet['_rank'] if (cadet and cadet.get('_rank')) else ('Admin' if is_admin else 'Cadet')
        cadet_name = cadet['name'] if (cadet and cadet.get('name')) else 'User'

        return jsonify({
            'success': True,
            'user': {
                'id': user_id,
                'email': email,
                'dli_number': dli,
                'is_admin': bool(is_admin),
                'rank': rank_name,
                'name': cadet_name
            }
        }), 201

    except mysql.connector.IntegrityError:
        conn.rollback()
        return jsonify({'error': 'An account with this email or DLI number already exists.'}), 409
    except Exception as e:
        conn.rollback()
        return jsonify({'error': f'Registration failed: {str(e)}'}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/auth/signin', methods=['POST'])
def auth_signin():
    """Sign in with email or DLI + password."""
    data = request.get_json() or {}
    identifier = (data.get('identifier') or '').strip()
    password = data.get('password') or ''

    if not identifier or not password:
        return jsonify({'error': 'Email/DLI and password are required.'}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        # Try email first, then DLI
        cursor.execute(
            "SELECT * FROM users WHERE email = %s OR dli_number = %s",
            (identifier.lower(), identifier.upper())
        )
        user = cursor.fetchone()

        if not user:
            return jsonify({'error': 'No account found with this email or DLI number.'}), 404

        if not check_password_hash(user['password_hash'], password):
            return jsonify({'error': 'Incorrect password.'}), 401

        # Set session
        session.permanent = True
        session['user_id'] = user['id']
        session['email'] = user['email']
        session['dli_number'] = user['dli_number']
        session['is_admin'] = bool(user['is_admin'])

        # Fetch cadet details if exists
        cursor.execute("SELECT name, _rank FROM cadets WHERE dli = %s LIMIT 1", (user['dli_number'],))
        cadet = cursor.fetchone()
        rank_name = cadet['_rank'] if (cadet and cadet.get('_rank')) else ('Admin' if user['is_admin'] else 'Cadet')
        cadet_name = cadet['name'] if (cadet and cadet.get('name')) else 'User'

        return jsonify({
            'success': True,
            'user': {
                'id': user['id'],
                'email': user['email'],
                'dli_number': user['dli_number'],
                'is_admin': bool(user['is_admin']),
                'rank': rank_name,
                'name': cadet_name
            }
        })

    finally:
        cursor.close()
        conn.close()


@app.route('/auth/logout', methods=['POST'])
def auth_logout():
    """Clear session."""
    session.clear()
    return jsonify({'success': True})


@app.route('/auth/me')
def auth_me():
    """Get current authenticated user info."""
    if 'user_id' not in session:
        return jsonify({'authenticated': False}), 200

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT name, _rank FROM cadets WHERE dli = %s LIMIT 1", (session.get('dli_number'),))
        cadet = cursor.fetchone()
        rank_name = cadet['_rank'] if (cadet and cadet.get('_rank')) else ('Admin' if session.get('is_admin') else 'Cadet')
        cadet_name = cadet['name'] if (cadet and cadet.get('name')) else 'User'
        
        return jsonify({
            'authenticated': True,
            'user': {
                'id': session['user_id'],
                'email': session.get('email'),
                'dli_number': session.get('dli_number'),
                'is_admin': session.get('is_admin', False),
                'rank': rank_name,
                'name': cadet_name
            }
        })
    except Exception as e:
        return jsonify({
            'authenticated': True,
            'user': {
                'id': session['user_id'],
                'email': session.get('email'),
                'dli_number': session.get('dli_number'),
                'is_admin': session.get('is_admin', False),
                'rank': 'Admin' if session.get('is_admin') else 'Cadet',
                'name': 'User'
            }
        })
    finally:
        cursor.close()
        conn.close()


@app.route('/auth/forgot-password', methods=['POST'])
def auth_forgot_password():
    """Generate and send a password reset code."""
    data = request.get_json() or {}
    email = (data.get('email') or '').strip().lower()

    if not email or not validate_email(email):
        return jsonify({'error': 'Please enter a valid email address.'}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT id, email FROM users WHERE email = %s", (email,))
        user = cursor.fetchone()

        if not user:
            # Don't reveal whether email exists — but per spec, we show a message
            return jsonify({'error': 'No account found with this email address.'}), 404

        # Generate 6-digit code
        reset_code = ''.join([str(secrets.randbelow(10)) for _ in range(6)])
        code_hash = hashlib.sha256(reset_code.encode()).hexdigest()
        expires_at = datetime.now() + timedelta(minutes=Config.RESET_TOKEN_EXPIRY_MINUTES)

        # Invalidate previous unused tokens for this user
        cursor.execute(
            "UPDATE password_reset_tokens SET used = 1 WHERE user_id = %s AND used = 0",
            (user['id'],)
        )

        # Store hashed code
        cursor.execute(
            "INSERT INTO password_reset_tokens (user_id, token_hash, expires_at) VALUES (%s, %s, %s)",
            (user['id'], code_hash, expires_at)
        )
        conn.commit()

        # Send email
        sent = send_reset_email(user['email'], reset_code)
        if not sent:
            return jsonify({'error': 'Failed to send reset email. Please try again.'}), 500

        return jsonify({
            'success': True,
            'message': 'Reset code sent to your email.',
            'email': email,
        })

    finally:
        cursor.close()
        conn.close()


@app.route('/auth/reset-password', methods=['POST'])
def auth_reset_password():
    """Verify reset code and set new password."""
    data = request.get_json() or {}
    email = (data.get('email') or '').strip().lower()
    code = (data.get('code') or '').strip()
    new_password = data.get('new_password') or ''
    confirm = data.get('confirm_password') or ''

    if not email or not code:
        return jsonify({'error': 'Email and reset code are required.'}), 400
    if not new_password or len(new_password) < 6:
        return jsonify({'error': 'Password must be at least 6 characters.'}), 400
    if new_password != confirm:
        return jsonify({'error': 'Passwords do not match.'}), 400

    code_hash = hashlib.sha256(code.encode()).hexdigest()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        # Find user
        cursor.execute("SELECT id FROM users WHERE email = %s", (email,))
        user = cursor.fetchone()
        if not user:
            return jsonify({'error': 'No account found with this email.'}), 404

        # Find valid token
        cursor.execute(
            """
            SELECT id, expires_at FROM password_reset_tokens
            WHERE user_id = %s AND token_hash = %s AND used = 0
            ORDER BY created_at DESC LIMIT 1
            """,
            (user['id'], code_hash)
        )
        token = cursor.fetchone()

        if not token:
            return jsonify({'error': 'Invalid or expired reset code.'}), 400

        if token['expires_at'] < datetime.now():
            cursor.execute("UPDATE password_reset_tokens SET used = 1 WHERE id = %s", (token['id'],))
            conn.commit()
            return jsonify({'error': 'Reset code has expired. Please request a new one.'}), 400

        # Update password
        pw_hash = generate_password_hash(new_password)
        cursor.execute("UPDATE users SET password_hash = %s WHERE id = %s", (pw_hash, user['id']))

        # Mark token as used
        cursor.execute("UPDATE password_reset_tokens SET used = 1 WHERE id = %s", (token['id'],))
        conn.commit()

        return jsonify({'success': True, 'message': 'Password has been reset successfully.'})

    finally:
        cursor.close()
        conn.close()


@app.route('/auth/make-admin', methods=['POST'])
@admin_required
def auth_make_admin():
    """Admin promotes another user to admin by DLI number."""
    data = request.get_json() or {}
    target_dli = (data.get('dli_number') or '').strip().upper()

    if not target_dli:
        return jsonify({'error': 'DLI number is required.'}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT id, email, dli_number, is_admin FROM users WHERE dli_number = %s", (target_dli,))
        target = cursor.fetchone()

        if not target:
            return jsonify({'error': 'No user found with this DLI number.'}), 404
        if target['is_admin']:
            return jsonify({'error': 'This user is already an admin.'}), 409

        cursor.execute("UPDATE users SET is_admin = 1 WHERE id = %s", (target['id'],))
        conn.commit()

        return jsonify({
            'success': True,
            'message': f"{target['email']} ({target['dli_number']}) has been promoted to admin."
        })

    finally:
        cursor.close()
        conn.close()


# ══════════════════════════════════════════════════════════════
#   CADET ATTENDANCE VIEW (for logged-in cadets)
# ══════════════════════════════════════════════════════════════

@app.route('/api/my-attendance')
@login_required
def api_my_attendance():
    """Get attendance data for the currently logged-in cadet."""
    dli = session.get('dli_number', '')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        # Find the cadet
        cursor.execute("SELECT id, name, dli, year, _rank FROM cadets WHERE dli = %s LIMIT 1", (dli,))
        cadet = cursor.fetchone()

        if not cadet:
            return jsonify({'error': 'Cadet profile not found for your DLI number.', 'cadet': None}), 200

        # Get all sessions with attendance status for this cadet
        cursor.execute(
            """
            SELECT cs.id, DATE_FORMAT(cs.date, '%%Y-%%m-%%d') AS date, cs.notes,
                   IFNULL(a.status, 'A') AS status,
                   MONTH(cs.date) AS month, YEAR(cs.date) AS cal_year
            FROM class_sessions cs
            LEFT JOIN attendance a ON a.session_id = cs.id AND a.cadet_id = %s
            ORDER BY cs.date DESC
            """,
            (cadet['id'],)
        )
        sessions = cursor.fetchall()

        # Compute monthly breakdown
        monthly = {}
        for s in sessions:
            key = f"{s['cal_year']}-{s['month']:02d}"
            if key not in monthly:
                monthly[key] = {'total': 0, 'present': 0, 'year': s['cal_year'], 'month': s['month']}
            monthly[key]['total'] += 1
            if s['status'] == 'P':
                monthly[key]['present'] += 1

        # Overall stats
        total = len(sessions)
        present = sum(1 for s in sessions if s['status'] == 'P')

        # Get session photos for sessions this cadet attended
        attended_ids = [s['id'] for s in sessions if s['status'] == 'P']
        photos = []
        if attended_ids:
            placeholders = ','.join(['%s'] * len(attended_ids))
            cursor.execute(
                f"""
                SELECT sp.id, sp.session_id, sp.filename, sp.original_name,
                       DATE_FORMAT(cs.date, '%%Y-%%m-%%d') AS session_date
                FROM session_photos sp
                JOIN class_sessions cs ON cs.id = sp.session_id
                WHERE sp.session_id IN ({placeholders})
                ORDER BY cs.date DESC
                LIMIT 20
                """,
                attended_ids
            )
            photo_rows = cursor.fetchall()
            for p in photo_rows:
                photos.append({
                    'id': p['id'],
                    'url': f"/static/images/session_photos/{p['session_id']}/{p['filename']}",
                    'date': p['session_date'],
                    'name': p['original_name'] or p['filename'],
                })

        return jsonify({
            'cadet': cadet,
            'total_sessions': total,
            'present': present,
            'percentage': round((present * 100 / total), 1) if total > 0 else 0,
            'sessions': sessions[:50],  # limit to recent 50
            'monthly': sorted(monthly.values(), key=lambda x: (x['year'], x['month']), reverse=True),
            'photos': photos,
        })

    finally:
        cursor.close()
        conn.close()


# ══════════════════════════════════════════════════════════════
#   EXISTING ROUTES (preserved + bug-fixed)
# ══════════════════════════════════════════════════════════════

@app.route('/api/ranks')
def api_ranks():
    session_label = request.args.get('session', None)
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id, label, is_current FROM rank_sessions ORDER BY label DESC")
    all_sessions = cursor.fetchall()
    if session_label:
        cursor.execute("SELECT id, label FROM rank_sessions WHERE label = %s", (session_label,))
    else:
        cursor.execute("SELECT id, label FROM rank_sessions WHERE is_current = 1 LIMIT 1")
    active_session = cursor.fetchone()
    if not active_session and all_sessions:
        active_session = all_sessions[0]
    rank_holders = []
    if active_session:
        cursor.execute(
            """
            SELECT rank_title, name, photo, rank_order, year_batch
            FROM rank_holders
            WHERE session_id = %s
            ORDER BY rank_order ASC, name ASC
            """,
            (active_session['id'],)
        )
        rank_holders = cursor.fetchall()
    cursor.close()
    conn.close()

    grouped = {}
    for holder in rank_holders:
        title = holder['rank_title']
        if title not in grouped:
            grouped[title] = []
        grouped[title].append(holder)

    return jsonify({
        'all_sessions': all_sessions,
        'active_session': active_session,
        'grouped': grouped
    })


# ── Achievements Routes ────────────────────────────────────────

@app.route('/api/achievements')
def api_get_achievements():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM achievements ORDER BY id DESC")
        rows = cursor.fetchall()
        return jsonify(rows)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/api/achievements/add', methods=['POST'])
@admin_required
def api_add_achievement():
    title = request.form.get('title', '').strip()
    description = request.form.get('description', '').strip()
    tag = request.form.get('tag', '').strip()
    
    if not title or not description or not tag:
        return jsonify({'error': 'Title, description, and tag are required.'}), 400
        
    file = request.files.get('image')
    if not file or not file.filename:
        return jsonify({'error': 'Image file is required.'}), 400
        
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in PHOTO_ALLOWED:
        return jsonify({'error': 'Invalid image type. Allowed: jpg, jpeg, png, webp, gif, heic'}), 400
        
    # Save file
    filename = f"{uuid.uuid4().hex}{ext}"
    upload_dir = get_achievements_dir()
    file.save(os.path.join(upload_dir, filename))
    image_path = f"/static/images/achievements/{filename}"
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            "INSERT INTO achievements (title, description, image_path, tag) VALUES (%s, %s, %s, %s)",
            (title, description, image_path, tag)
        )
        conn.commit()
        ach_id = cursor.lastrowid
        return jsonify({
            'success': True,
            'achievement': {
                'id': ach_id,
                'title': title,
                'description': description,
                'image_path': image_path,
                'tag': tag
            }
        }), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route("/api/cadets")
def api_cadets():
    year = request.args.get("year")
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    if year == 'all':
        cursor.execute("SELECT * FROM cadets ORDER BY name")
    elif year:
        cursor.execute("SELECT * FROM cadets WHERE year = %s ORDER BY name", (year,))
    else:
        cursor.execute("SELECT * FROM cadets ORDER BY name")
    cadets = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(cadets)


# ── Attendance Routes ─────────────────────────────────────────

@app.route('/attendance/cadets')
def attendance_cadets():
    year = request.args.get('year', '')
    if year not in ('1', '2', '3'):
        return jsonify({'error': 'Invalid year'}), 400
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        "SELECT id, name, dli, year, _rank FROM cadets WHERE year=%s ORDER BY name",
        (year,)
    )
    cadets = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(cadets)


@app.route('/attendance/sessions')
def attendance_sessions():
    year = request.args.get('year', '')
    month = request.args.get('month', '')
    if not year or not month:
        return jsonify([])
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT cs.id,
               DATE_FORMAT(cs.date,'%%Y-%%m-%%d') AS date,
               cs.notes,
               COUNT(sp.id) AS photo_count
        FROM class_sessions cs
        LEFT JOIN session_photos sp ON sp.session_id = cs.id
        WHERE YEAR(cs.date)=%s AND MONTH(cs.date)=%s
        GROUP BY cs.id, cs.date, cs.notes
        ORDER BY cs.date
        """,
        (year, month)
    )
    sessions = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(sessions)


@app.route('/attendance/sessions/add', methods=['POST'])
@admin_required
def attendance_add_session():
    data = request.get_json()
    date = data.get('date', '').strip()
    notes = data.get('notes', '').strip()
    if not date:
        return jsonify({'error': 'Date required'}), 400
    try:
        datetime.strptime(date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            "INSERT INTO class_sessions (date, notes) VALUES (%s, %s)",
            (date, notes or None)
        )
        conn.commit()
        session_id = cursor.lastrowid
    except mysql.connector.IntegrityError:
        cursor.close()
        conn.close()
        return jsonify({'error': 'Session for this date already exists'}), 409
    except Exception as e:
        cursor.close()
        conn.close()
        return jsonify({'error': str(e)}), 500

    cursor.close()
    conn.close()
    return jsonify({'id': session_id, 'date': date, 'notes': notes})


@app.route('/attendance/sessions/<int:session_id>/photos')
def get_session_photos(session_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        "SELECT id, filename, original_name, uploaded_at FROM session_photos WHERE session_id = %s ORDER BY uploaded_at ASC",
        (session_id,)
    )
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    photos = []
    for row in rows:
        url = f"/static/images/session_photos/{session_id}/{row['filename']}"
        photos.append({
            'id': row['id'],
            'filename': row['original_name'] or row['filename'],
            'url': url,
            'uploaded_at': str(row['uploaded_at'])
        })
    return jsonify({'photos': photos})


@app.route('/attendance/sessions/<int:session_id>/photos/upload', methods=['POST'])
@admin_required
def upload_session_photos(session_id):
    files = request.files.getlist('photos')
    if not files:
        return jsonify({'error': 'No files uploaded'}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id FROM class_sessions WHERE id = %s", (session_id,))
    if not cursor.fetchone():
        cursor.close()
        conn.close()
        return jsonify({'error': 'Session not found'}), 404

    photo_dir = get_session_photos_dir(session_id)
    uploaded = 0

    for f in files:
        if not f or not f.filename:
            continue
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in PHOTO_ALLOWED:
            continue
        original_name = secure_filename(f.filename)
        stored_name = f"{uuid.uuid4().hex}{ext}"
        f.save(os.path.join(photo_dir, stored_name))
        cursor.execute(
            "INSERT INTO session_photos (session_id, filename, original_name) VALUES (%s, %s, %s)",
            (session_id, stored_name, original_name)
        )
        uploaded += 1

    conn.commit()
    cursor.close()
    conn.close()

    if uploaded == 0:
        return jsonify({'error': 'No valid image files found'}), 400
    return jsonify({'uploaded': uploaded})


@app.route('/attendance/sessions/photos/<int:photo_id>', methods=['DELETE'])
@admin_required
def delete_session_photo(photo_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM session_photos WHERE id = %s", (photo_id,))
    photo = cursor.fetchone()
    if not photo:
        cursor.close()
        conn.close()
        return jsonify({'error': 'Photo not found'}), 404

    photo_path = os.path.join(
        app.root_path, 'static', 'images', 'session_photos',
        str(photo['session_id']), photo['filename']
    )
    try:
        if os.path.exists(photo_path):
            os.remove(photo_path)
    except Exception:
        pass

    cursor.execute("DELETE FROM session_photos WHERE id = %s", (photo_id,))
    conn.commit()
    cursor.close()
    conn.close()
    return jsonify({'deleted': True})


@app.route('/attendance/sessions/<int:session_id>', methods=['DELETE'])
@admin_required
def delete_session(session_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT id FROM class_sessions WHERE id = %s", (session_id,))
    if not cursor.fetchone():
        cursor.close()
        conn.close()
        return jsonify({'error': 'Session not found'}), 404

    try:
        cursor.execute("DELETE FROM attendance WHERE session_id = %s", (session_id,))

        cursor.execute("SELECT filename FROM session_photos WHERE session_id = %s", (session_id,))
        photos = cursor.fetchall()
        for p in photos:
            photo_path = os.path.join(
                app.root_path, 'static', 'images', 'session_photos',
                str(session_id), p['filename']
            )
            try:
                if os.path.exists(photo_path):
                    os.remove(photo_path)
            except Exception:
                pass
        cursor.execute("DELETE FROM session_photos WHERE session_id = %s", (session_id,))

        photo_dir = os.path.join(app.root_path, 'static', 'images', 'session_photos', str(session_id))
        try:
            if os.path.exists(photo_dir):
                import shutil
                shutil.rmtree(photo_dir)
        except Exception:
            pass

        cursor.execute("DELETE FROM class_sessions WHERE id = %s", (session_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'error': str(e)}), 500

    cursor.close()
    conn.close()
    return jsonify({'deleted': True})


@app.route('/attendance/records')
def attendance_records():
    session_id = request.args.get('session_id', '')
    cadet_year = request.args.get('year', '')
    if not session_id or cadet_year not in ('1', '2', '3'):
        return jsonify([])
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT c.id, c.name, c.dli, c.year, c._rank,
               IFNULL(a.status, 'A') AS status
        FROM cadets c
        LEFT JOIN attendance a
          ON a.cadet_id = c.id AND a.session_id = %s
        WHERE c.year = %s
        ORDER BY c.name
        """,
        (session_id, cadet_year)
    )
    records = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(records)


@app.route('/attendance/mark', methods=['POST'])
@admin_required
def attendance_mark():
    data = request.get_json()
    session_id = data.get('session_id')
    records = data.get('records', [])
    if not session_id or not records:
        return jsonify({'error': 'Missing data'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        for rec in records:
            cursor.execute(
                """
                INSERT INTO attendance (cadet_id, session_id, status)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE status = VALUES(status)
                """,
                (rec['cadet_id'], session_id, rec['status'])
            )
        conn.commit()
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'error': str(e)}), 500

    cursor.close()
    conn.close()
    return jsonify({'saved': len(records)})


@app.route('/attendance/analytics')
def attendance_analytics():
    cal_year = request.args.get('year', '')
    month = request.args.get('month', '')
    cadet_year = request.args.get('cadet_year', '')
    if not cal_year or not month or cadet_year not in ('1', '2', '3'):
        return jsonify({'error': 'Missing params'}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute(
        "SELECT COUNT(*) AS total FROM class_sessions WHERE YEAR(date)=%s AND MONTH(date)=%s",
        (cal_year, month)
    )
    total_sessions = cursor.fetchone()['total']

    cursor.execute(
        """
        SELECT c.id, c.name, c.dli, c._rank,
               COUNT(a.id)                                   AS attended,
               %s                                            AS total,
               ROUND(COUNT(a.id) * 100.0 / NULLIF(%s,0), 1) AS pct
        FROM cadets c
        LEFT JOIN attendance a
          ON  a.cadet_id  = c.id
          AND a.status    = 'P'
          AND a.session_id IN (
              SELECT id FROM class_sessions
              WHERE YEAR(date)=%s AND MONTH(date)=%s
          )
        WHERE c.year = %s
        GROUP BY c.id, c.name, c.dli, c._rank
        ORDER BY pct DESC, c.name
        """,
        (total_sessions, total_sessions, cal_year, month, cadet_year)
    )
    cadets = cursor.fetchall()
    cursor.close()
    conn.close()

    return jsonify({
        'total_sessions': total_sessions,
        'cadets': cadets,
    })


# ── Camps Routes ──────────────────────────────────────────────

@app.route('/api/camps')
def api_camps():
    camps_dir = get_camps_dir()
    camps_list = []
    try:
        entries = sorted(os.listdir(camps_dir))
    except FileNotFoundError:
        entries = []
    for entry in entries:
        entry_path = os.path.join(camps_dir, entry)
        if not os.path.isdir(entry_path):
            continue
        images = list_images(entry_path)
        camps_list.append({
            'name': entry.upper().replace('_', ' '),
            'folder': entry,
            'cover_file': images[0] if images else None,
            'count': len(images),
        })
    return jsonify(camps_list)


@app.route('/api/camp/<camp_name>')
def api_camp_gallery(camp_name):
    safe_name = ''.join(c for c in camp_name if c.isalnum() or c in '-_')
    if not safe_name:
        return jsonify({'error': 'Invalid camp name'}), 400
    camp_path = os.path.join(get_camps_dir(), safe_name)
    if not os.path.isdir(camp_path):
        return jsonify({'error': 'Camp not found'}), 404
    image_files = list_images(camp_path)
    return jsonify({
        'camp_name': safe_name.upper().replace('_', ' '),
        'camp_folder': safe_name,
        'images': image_files
    })


# ── Nominal Roll Routes ──────────────────────────────────────

@app.route('/admin/nominal-roll/upload', methods=['POST'])
@admin_required
def upload_excel():
    """Upload Excel, return columns + cadet list as JSON."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    f = request.files['file']
    if f.filename == '':
        return jsonify({'error': 'Empty filename'}), 400

    if not f.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
        return jsonify({'error': 'Invalid file type. Upload .xlsx, .xls or .csv'}), 400

    try:
        if f.filename.lower().endswith('.csv'):
            df = pd.read_csv(f)
        else:
            df = pd.read_excel(f)

        if df.empty:
            return jsonify({'error': 'File is empty or has no data rows'}), 400

        df = df.fillna('')
        columns = list(df.columns)
        records = df.to_dict(orient='records')

        tmp = tempfile.NamedTemporaryFile(
            delete=False, suffix='.json',
            dir=tempfile.gettempdir(),
            prefix='ncc_roll_'
        )
        tmp.write(json.dumps({'columns': columns, 'records': records}).encode())
        tmp.close()

        return jsonify({
            'success': True,
            'columns': columns,
            'cadets': records,
            'temp_file': os.path.basename(tmp.name),
            'total': len(records)
        })

    except Exception as e:
        return jsonify({'error': f'Failed to read file: {str(e)}'}), 500


@app.route('/admin/nominal-roll/generate', methods=['POST'])
@admin_required
def generate_nominal_roll():
    """Generate sorted, filtered Excel from selected cadets + columns."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data received'}), 400

    temp_file = data.get('temp_file')
    selected_ids = data.get('selected_ids', [])
    selected_cols = data.get('columns', [])
    sort_by = data.get('sort_by', 'none')
    serial_col = data.get('serial_col', True)

    if not temp_file:
        return jsonify({'error': 'Session expired. Please re-upload the file.'}), 400

    tmp_path = os.path.join(tempfile.gettempdir(), temp_file)
    if not os.path.exists(tmp_path):
        return jsonify({'error': 'Temp file not found. Please re-upload.'}), 400

    try:
        with open(tmp_path, 'r') as fp:
            payload = json.load(fp)

        all_records = payload['records']

        if selected_ids:
            records = [all_records[i] for i in selected_ids if i < len(all_records)]
        else:
            return jsonify({'error': 'No cadets selected'}), 400

        df = pd.DataFrame(records)

        if selected_cols:
            valid_cols = [c for c in selected_cols if c in df.columns]
            if not valid_cols:
                return jsonify({'error': 'None of the selected columns exist in the data'}), 400
            df = df[valid_cols]

        if sort_by == 'rank':
            rank_col = next((c for c in df.columns if 'rank' in c.lower()), None)
            if rank_col:
                df['_rank_priority'] = df[rank_col].apply(get_rank_priority)
                df = df.sort_values('_rank_priority').drop(columns=['_rank_priority'])

        elif sort_by == 'dli':
            dli_col = next((c for c in df.columns if 'dli' in c.lower()), None)
            if dli_col:
                df['_dli_year'] = df[dli_col].apply(extract_dli_year)
                df = df.sort_values('_dli_year').drop(columns=['_dli_year'])

        if serial_col:
            df.insert(0, 'S.No', range(1, len(df) + 1))

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Nominal Roll')

            worksheet = writer.sheets['Nominal Roll']
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='1E3A1E', end_color='1E3A1E', fill_type='solid')
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                col_letter = cell.column_letter
                worksheet.column_dimensions[col_letter].width = max(15, len(str(cell.value)) + 4)

            alt_fill = PatternFill(start_color='F0F4EC', end_color='F0F4EC', fill_type='solid')
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if row_idx % 2 == 0:
                        cell.fill = alt_fill

            worksheet.freeze_panes = 'A2'

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='NSUT_NCC_Nominal_Roll.xlsx'
        )

    except Exception as e:
        return jsonify({'error': f'Generation failed: {str(e)}'}), 500

    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass


# ── SPA Catch-All ─────────────────────────────────────────────

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def catch_all(path):
    file_path = os.path.join(app.static_folder, path)
    if path and os.path.exists(file_path):
        return send_from_directory(app.static_folder, path)
    return send_file(os.path.join(app.static_folder, 'index.html'))


# ── Entry Point ───────────────────────────────────────────────

if __name__ == '__main__':
    ensure_tables()
    app.run(debug=True)